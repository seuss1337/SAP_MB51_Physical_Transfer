# -*- coding: utf-8 -*-
"""Untitled0.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1eekHIPklDGWRjt0xOyQp_oi68XDsSXTB
"""

import pandas as pd
import numpy as np
import openpyxl as xl
import os
import shutil

movements_df = pd.read_excel(r"/content/drive/My Drive/Colab Notebooks/EXPORTmb51.XLSX")

doc_nums_raw = movements_df["Material Document"].tolist()
doc_nums = []  # list of doc nums from mb51
for i in doc_nums_raw:
    if str(i)[:3] == "490" and (i in doc_nums) == False:
        doc_nums.append(np.int64(i))
temp = movements_df.loc[movements_df["Material Document"] == doc_nums[0]]


for x in range(len(doc_nums)):
    current_df = movements_df.loc[movements_df["Material Document"] == doc_nums[x]]
    for index, row in current_df.iterrows():
        print(row["Material"], row["Qty in Un. of Entry"])
        print(type(row["Material"]), type(row["Qty in Un. of Entry"]))
    print()

# copy file to new file, copy polish sheet to new sheet
# have to copy to new file because openpyxl won't support copying between workbooks

source_dir = "/content/drive/My Drive/Colab Notebooks/MASTER.xlsx"
destination_dir = "/content/drive/My Drive/Colab Notebooks/TEST.xlsx"
shutil.copy(source_dir, destination_dir)
# coppy part
wb = xl.load_workbook(destination_dir)
ws = wb.worksheets[0]


# for i in range(12, 65):
#   if ws.cell(row = i, column = 1).value == "CM9048SASEBL25-EA":
#     ws.cell(row = i, column = 3, value = i)
wb.save(str(destination_dir))

# Skeleton final working
in_sheet = False
# loop through doc num list
for x in range(len(doc_nums)):
    # for x in range(1):

    # ws = wb.worksheets[0]
    # source = wb.active
    # wb.copy_worksheet(ws)
    # wb.worksheets[5+x].title = str(doc_nums[x])
    # ws =  wb.worksheets[5+x]
    wb.copy_worksheet(ws)
    wb.worksheets[5 + x].title = str(doc_nums[x])
    ws = wb.worksheets[5 + x]

    # temp code to make new worksheet named doc number
    not_in_list_counter = 55
    current_df = movements_df.loc[movements_df["Material Document"] == doc_nums[x]]

    # ws.cell(row = 2, column = 5, value = current_df.at[2, 'Storage Location']) #Job Number
    # ws.cell(row = 3, column = 5, value = str(current_df.at[2, 'Posting Date'])) #DATE
    ws.cell(row=3, column=5, value="testing")  # DATE
    ws.cell(row=4, column=5, value=str(doc_nums[x]))  # Doc Number

    # go through dataframe containing document
    for index, row in current_df.iterrows():
        if row["Qty in Un. of Entry"] < 0:
            ws.cell(row=7, column=6, value="x")
        else:
            ws.cell(row=7, column=3, value="x")
        # loop through destination file
        for i in range(12, 65):
            if ws.cell(row=i, column=1).value == row["Material"]:
                ws.cell(row=i, column=3, value=row["Qty in Un. of Entry"])
                in_sheet = True
            if ws.cell(row=i, column=4).value == row["Material"]:
                ws.cell(row=i, column=6, value=row["Qty in Un. of Entry"])
                in_sheet = True

        if in_sheet == False:
            ws.cell(row=not_in_list_counter, column=4, value=row["Material"])
            ws.cell(row=not_in_list_counter, column=6, value=row["Qty in Un. of Entry"])
            not_in_list_counter += 1
        in_sheet = False
    print(doc_nums[x])
    print(x)
wb.save(str(destination_dir))
